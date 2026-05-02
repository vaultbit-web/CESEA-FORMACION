"""
CESEA Formación - build-migrations.py
========================================
Convierte los Excels en assets/ a JSONs en migrations/ que el script de
Node consume para hacer la migración a Supabase.

Salida:
  migrations/data-formadores.json   (35 formadores reales del Excel)
  migrations/data-cursos.json       (113 cursos del Excel)

No requiere parámetros. Re-ejecutable.
"""
import openpyxl, json, unicodedata, re, sys, os

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ASSETS = os.path.join(ROOT, 'assets')
OUT = os.path.join(ROOT, 'migrations')

def slug_name(s):
    s = (s or '').lower().strip()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^a-z0-9]', ' ', s)
    return ' '.join(s.split())

def norm_grupo(g):
    g = (g or '').upper().strip()
    g_ascii = g.replace('Á','A').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ú','U')
    if 'TECNICA' in g_ascii: return 'Área técnica'
    if 'ACP' in g_ascii: return 'ACP y Modelo de Atención'
    if 'AUTOCUIDADO' in g_ascii: return 'Área de autocuidados'
    if 'COMPETENCIA' in g_ascii: return 'Competencias'
    return 'Área técnica'

def norm_familia(f):
    f = (f or '').upper().strip()
    M = {
        'ACOMPAÑAMIENTO':'Acompañamiento','ACP':'ACP',
        'ATENCIÓN BUCODENTAL':'Atención bucodental',
        'ATENCIÓN MÉDICA Y CUIDADOS':'Atención médica y cuidados',
        'BIENESTAR EMOCIONAL Y AUTOCUIDADO':'Bienestar emocional y autocuidado',
        'COMPETENCIAS COMUNICATIVAS':'Competencias comunicativas',
        'CUIDADOS EN DOMICILIO':'Cuidados en domicilio',
        'FARMACIA':'Farmacia',
        'LIDERAZGO Y TRABAJO EN EQUIPO':'Liderazgo y trabajo en equipo',
        'MOVILIZACIONES':'Movilizaciones','NUTRICIÓN':'Nutrición',
        'PERSPECTIVA DE GÉNERO':'Perspectiva de género',
        'SEGURIDAD LABORAL Y PROTOCOLOS':'Seguridad laboral y protocolos',
        'VOLUNTARIADO':'Voluntariado',
        'ÉTICA':'Ética profesional','ÉTICA PROFESIONAL':'Ética profesional',
    }
    return M.get(f, f.title())

def first_objective(text):
    if not text: return ''
    text = text.replace('\r','').strip()
    m = re.match(r'^\s*1[\.\)]\s*(.+?)(?:\n\n|\n\s*2[\.\)]|$)', text, re.S)
    if m:
        first = m.group(1).strip().split('\n')[0]
        return first[:240].rstrip(' .,;') + ('...' if len(first) > 240 else '.')
    return text[:240].split('\n')[0]

# ── 1. Formadores ──────────────────────────────────────────────────────────
print('Loading Datos_Formadores.xlsx...')
wb = openpyxl.load_workbook(os.path.join(ASSETS, 'Datos_Formadores.xlsx'), data_only=True)
ws = wb['Sheet1']
formadores = []
for row in ws.iter_rows(values_only=True, min_row=2):
    if not row[0]: continue
    formadores.append({
        'idExterno': int(row[0]),
        'name': (row[1] or '').strip(),
        'cif': row[2],
        'nColegiado': row[3],
        'email': (row[4] or '').strip(),
        'phone': str(row[5]) if row[5] else '',
        'poblacion': (row[6] or '').strip(),
    })
print(f'  ->{len(formadores)} formadores extraídos')

# ── 2. Cursos ──────────────────────────────────────────────────────────────
print('Loading Formadores_Curso.xlsx...')
wb2 = openpyxl.load_workbook(os.path.join(ASSETS, 'Formadores_Curso.xlsx'), data_only=True)
ws2 = wb2['Sheet1']

cursos = []
current = None
for row in ws2.iter_rows(values_only=True, min_row=2):
    codigo = row[0]
    if codigo:
        if current: cursos.append(current)
        current = {
            'codigo': str(codigo).strip(),
            'denominacion': (row[1] or '').strip(),
            'grupo': (row[2] or '').strip(),
            'familia': (row[3] or '').strip(),
            'contenidos': (row[4] or '').strip(),
            'objetivos': (row[5] or '').strip(),
            'profesionales': []
        }
    if current and (row[6] or row[7]):
        current['profesionales'].append({
            'ciudad': (row[6] or '').strip(),
            'profesional': (row[7] or '').strip(),
            'tipoContrato': row[8],
        })
if current: cursos.append(current)
print(f'  ->{len(cursos)} cursos únicos extraídos')

# ── 3. Cruce formador↔curso por nombre ─────────────────────────────────────
name_to_idext = {}     # slug -> idExterno (int) for known formadores
for f in formadores:
    name_to_idext[slug_name(f['name'])] = f['idExterno']

# Identificar profesionales del Excel de cursos que NO están en formadores
extras = {}
for c in cursos:
    for p in c['profesionales']:
        slug = slug_name(p['profesional'])
        if not slug or slug in name_to_idext: continue
        if slug not in extras:
            extras[slug] = {
                'name': p['profesional'].strip(),
                'poblacion': (p['ciudad'] or '').strip(),
            }

# Asignar IDs sintéticos negativos a los extras (para distinguirlos del Excel real)
extra_list = []
for i, (slug, info) in enumerate(extras.items(), start=1):
    fake_id = -i  # extras llevan idExterno negativo (-1, -2, ...)
    name_to_idext[slug] = fake_id
    extra_list.append({**info, 'idExterno': fake_id, 'cif': None, 'nColegiado': None, 'email': '', 'phone': ''})

print(f'  ->{len(extra_list)} extras encontrados en cursos pero no en formadores')

# Añadir extras a la lista de formadores
all_formadores = formadores + extra_list

# Construir cursos finales con cross-reference
final_cursos = []
for i, c in enumerate(cursos):
    fext_ids = []
    cities = []
    for p in c['profesionales']:
        slug = slug_name(p['profesional'])
        if slug and slug in name_to_idext:
            fid = name_to_idext[slug]
            if fid not in fext_ids: fext_ids.append(fid)
        if p['ciudad'] and p['ciudad'] not in cities: cities.append(p['ciudad'])
    final_cursos.append({
        'order': i + 1,
        'codigoInterno': c['codigo'],
        'title': c['denominacion'],
        'subtitle': first_objective(c['objetivos']),
        'area': norm_grupo(c['grupo']),
        'category': norm_familia(c['familia']),
        'contenidos': c['contenidos'],
        'objetivos': c['objetivos'],
        'tipoContrato': c['profesionales'][0]['tipoContrato'] if c['profesionales'] else None,
        'ciudades': cities,
        'formadoresIdExternos': fext_ids,   # int[] — el JS de migración resuelve a UUID
        'numImparticiones': 1,
        'modality': 'Presencial',
        'hours': 8,
        'level': 'Intermedio',
        'status': 'available',
    })

# ── 4. Volcado ─────────────────────────────────────────────────────────────
os.makedirs(OUT, exist_ok=True)
with open(os.path.join(OUT, 'data-formadores.json'), 'w', encoding='utf-8') as f:
    json.dump(all_formadores, f, ensure_ascii=False, indent=2)
with open(os.path.join(OUT, 'data-cursos.json'), 'w', encoding='utf-8') as f:
    json.dump(final_cursos, f, ensure_ascii=False, indent=2)

print()
print('OK Generados:')
print(f'  migrations/data-formadores.json ({len(all_formadores)} registros)')
print(f'  migrations/data-cursos.json     ({len(final_cursos)} registros)')
